#!/usr/bin/env python3
"""배치별 raw JSON 파일들을 병합·중복제거하고 표준 xlsx로 저장.

사용법:
  python merge_and_save.py \
      --raw-glob "output/naver.com/<channel>_<id>/_raw_batch*.json" \
      --output   "output/naver.com/<channel>_<id>/<slug>_<YYYYMMDD_HHMMSS>.xlsx"
"""
import argparse
import glob
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


COLUMNS = [
    ('id', '리뷰ID'),
    ('date', '작성일시'),
    ('score', '평점'),
    ('writer', '작성자(마스킹)'),
    ('option', '옵션'),
    ('content', '본문'),
    ('reviewContentClassType', '리뷰형식'),
    ('attachCount', '첨부수'),
    ('firstAttachUrl', '첫첨부URL'),
    ('repurchase', '재구매'),
    ('reviewType', '리뷰타입'),
    ('reviewServiceType', '서비스타입'),
    ('productNo', '상품번호'),
    ('productOrderNo', '주문번호'),
]
WIDTHS = {'리뷰ID': 12, '작성일시': 16, '평점': 6, '작성자(마스킹)': 14, '옵션': 18,
          '본문': 60, '리뷰형식': 10, '첨부수': 8, '첫첨부URL': 30, '재구매': 8,
          '리뷰타입': 12, '서비스타입': 12, '상품번호': 14, '주문번호': 18}


def load_items(paths: list[str]) -> list[dict]:
    items = []
    for p in paths:
        d = json.load(open(p, encoding='utf-8'))
        # agent-browser eval --json wraps result in {success, data: {origin, result}}
        if isinstance(d, dict) and 'data' in d and 'result' in d.get('data', {}):
            d = d['data']['result']
        if isinstance(d, dict) and 'items' in d:
            items.extend(d['items'])
        elif isinstance(d, list):
            items.extend(d)
        else:
            raise ValueError(f'Unrecognized raw shape in {p}')
    return items


def dedupe(items: list[dict]) -> list[dict]:
    seen = set()
    out = []
    for r in items:
        if r['id'] in seen:
            continue
        seen.add(r['id'])
        out.append(r)
    return out


def quality_summary(items: list[dict]) -> dict:
    n = len(items)
    if n == 0:
        return {'rows': 0}
    null_rate = {}
    for fld, _ in COLUMNS:
        nulls = sum(1 for r in items if r.get(fld) in (None, '', 0) and not isinstance(r.get(fld), bool))
        null_rate[fld] = round(nulls / n * 100, 1)
    score_dist = {}
    for r in items:
        score_dist[r.get('score')] = score_dist.get(r.get('score'), 0) + 1
    dates = sorted(r['date'][:10] for r in items if r.get('date'))
    return {
        'rows': n,
        'null_rate_pct': null_rate,
        'score_dist': dict(sorted(score_dist.items())),
        'date_range': [dates[0], dates[-1]] if dates else None,
    }


def write_xlsx(items: list[dict], out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = '리뷰'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2E5BBA')
    for ci, (_, label) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=ci, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
    for ri, item in enumerate(items, start=2):
        for ci, (key, _) in enumerate(COLUMNS, start=1):
            v = item.get(key)
            if key == 'date' and v:
                try:
                    dt = datetime.fromisoformat(v.replace('Z', '+00:00'))
                    v = dt.strftime('%Y-%m-%d %H:%M')
                except Exception:
                    pass
            if isinstance(v, bool):
                v = 'Y' if v else 'N'
            ws.cell(row=ri, column=ci, value=v)
    for ci, (_, label) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = WIDTHS.get(label, 14)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(COLUMNS)).column_letter}{len(items)+1}"
    content_ci = next(i for i, (k, _) in enumerate(COLUMNS, start=1) if k == 'content')
    for r in range(2, len(items) + 2):
        ws.cell(row=r, column=content_ci).alignment = Alignment(wrap_text=True, vertical='top')
    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--raw-glob', required=True, help='glob for raw batch JSONs')
    ap.add_argument('--output', required=True, help='output xlsx path')
    ap.add_argument('--also-json', help='also write merged json to this path')
    args = ap.parse_args()

    paths = sorted(glob.glob(args.raw_glob))
    if not paths:
        raise SystemExit(f'no files match: {args.raw_glob}')
    items = load_items(paths)
    deduped = dedupe(items)
    print(f'loaded={len(items)} dedup={len(deduped)} dupes_removed={len(items) - len(deduped)}')

    summary = quality_summary(deduped)
    print(json.dumps(summary, ensure_ascii=False, indent=2))

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    write_xlsx(deduped, out)
    print(f'saved: {out}')

    if args.also_json:
        Path(args.also_json).parent.mkdir(parents=True, exist_ok=True)
        json.dump(deduped, open(args.also_json, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
        print(f'json: {args.also_json}')


if __name__ == '__main__':
    main()
